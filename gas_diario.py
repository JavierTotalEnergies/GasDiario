import os
import requests
import pandas as pd
from datetime import datetime
import urllib3
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Desactivar advertencias de seguridad por el verify=False
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

commodities = [
    "NATURAL_GAS_GBP", 
    "BRENT_CRUDE_USD", 
    "JKM_LNG_USD", 
    "EU_CARBON_EUR"
]
api_key = "Token ee710707717f5ce9e3f22dbed30b8a2c2e068529948b4b203c5ddb1d8e624745"
headers = {"Authorization": api_key}

# ESTE ES EL NOMBRE FIJO (sin la fecha en el nombre)
nombre_archivo = "Historico_OilPrice.xlsx"  

datos_recolectados = []

print("Iniciando descarga de datos...")

for commodity in commodities:
    url = f"https://api.oilpriceapi.com/v1/commodities/{commodity}"
    
    try:
        response = requests.get(url, headers=headers, verify=False)
        response.raise_for_status() 
        
        data = response.json().get("data", {})
        
        if data:
            codigo = data.get("code")
            precio = data.get("current_price", {}).get("value")
            moneda = data.get("current_price", {}).get("currency")
            actualizacion = data.get("current_price", {}).get("last_updated")
            
            datos_recolectados.append({
                "Codigo": codigo,
                "Precio": precio,
                "Moneda": moneda,
                "Ultima_Actualizacion": actualizacion,
                "Fecha_Ejecucion": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            print(f"Datos obtenidos para: {commodity}")
            
    except requests.exceptions.RequestException as e:
        print(f"Error al consultar {commodity}: {e}")

# ----------------- LA NUEVA LÓGICA DE ACTUALIZACIÓN -----------------
if datos_recolectados:
    df_nuevo = pd.DataFrame(datos_recolectados)
    
    # Comprobar si el archivo de historial ya existe en la misma carpeta
    if os.path.exists(nombre_archivo):
        print(f"📂 Archivo existente detectado. Añadiendo nuevos datos...")
        # Leer la información pasada
        df_existente = pd.read_excel(nombre_archivo)
        # Combinar (concatenar) los datos pasados con los de hoy
        df_final = pd.concat([df_existente, df_nuevo], ignore_index=True)
    else:
        print(f"Creando nuevo archivo histórico por primera vez...")
        # Si no existe, usamos sólo los datos de hoy
        df_final = df_nuevo
    
    # Guardar todo el consolidado temporalmente en Excel
    df_final.to_excel(nombre_archivo, index=False)
    
    # --- BLOQUE OPCIONAL DE ESTILO VISUAL CON OPENPYXL ---
    # Esto asegura que el archivo mantenga siempre un formato limpio, bordes y el ancho correcto.
    try:
        wb = load_workbook(nombre_archivo)
        ws = wb.active
        ws.title = "Histórico Precios"
        
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        
        # Estilizar Cabecera
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Formateo básico de columnas (precio como número, ajuste de texto)
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for c_idx, cell in enumerate(row, start=1):
                if c_idx == 2:  # La columna de precio
                    cell.number_format = '#,##0.00'
        
        # Ajustar ancho de las columnas automáticamente
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
            
        wb.save(nombre_archivo)
        print(f"Datos guardados en: {nombre_archivo}")
        
    except Exception as e:
        print(f"Alerta al aplicar estilos visuales: {e}")

else:
    print("\nNo se pudieron obtener datos.")